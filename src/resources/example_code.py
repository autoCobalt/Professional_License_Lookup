import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict

def read_named_table_to_list_dict(file_path: str, table_name: str) -> List[Dict[str, str]]:
    # Load the workbook and the sheet
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Find the named table
    table = None
    for tbl in ws.tables.values():
        if tbl.name == table_name:
            table = tbl
            break

    if table is None:
        raise ValueError(f"Table '{table_name}' not found in the Excel file.")

    # Extract the table data
    data = ws[table.ref]

    # Convert the data to a list of dictionaries
    rows = list(data)
    headers = [cell.value for cell in rows[0]]
    list_dict = [
        {headers[i]: str(cell.value) for i, cell in enumerate(row)}
        for row in rows[1:]
    ]

    return list_dict

# Example usage
file_path = 'path/to/your/excel_file.xlsx'
table_name = 'emplid_lic_type'
data_list = read_named_table_to_list_dict(file_path, table_name)
print(data_list)