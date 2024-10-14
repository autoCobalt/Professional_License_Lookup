#1st party pre-installed libraries
import logging
import os
import tkinter as tk
from tkinter import filedialog
from typing import Dict, List, Union

# 3rd party libraries
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def load_emplid_data(file_path: Union[str, None] = None, dir_path: str = os.path.dirname(os.path.realpath(__file__)), sheet_name: str = 'Search_Request', table_name: str = 'emplid_lic_type') -> List[Dict[str, str]]:
    wb = read_excel_file(file_path=file_path,dir_path= dir_path)

    #Check if sheet exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")

    ws = wb[sheet_name]
    
    # Find sheet_name table
    table = None
    for tbl in ws.tables.values():
        if tbl.name == table_name:
            table = tbl
            break

    if table is None:
        raise ValueError(f"Table '{table_name}' not found in the Excel file.")
    
    data = ws[table.ref]

    rows = list(data)
    headers = [cell.value for cell in rows[0]]
    results = [
        {headers[i]: str(cell.value) for i, cell in enumerate(row) if cell.value} 
        for row in rows[1:] 
        if any(cell.value for cell in row)
    ]
    return results



def read_excel_file(file_path: Union[str, None] = None, dir_path: str = os.path.dirname(os.path.realpath(__file__))) -> openpyxl.workbook.Workbook:
    if file_path is None:
        file_path = __select_file(dir_path=dir_path)
    if file_path is None or file_path == '':
        raise ValueError("No file selected.")

    extension_type = os.path.splitext(file_path)[1]

    return load_workbook(filename=file_path, data_only=True)

def __select_file(dir_path: str = None) -> str:
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select an Excel/CSV file",
        filetypes=(("Excel file", "*.xlsx"), ("", "")),
        initialdir=dir_path,
        multiple=False
    )
    return file_path